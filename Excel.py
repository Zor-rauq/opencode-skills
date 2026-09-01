#!/usr/bin/env python3
"""
Reproduit les fusions verticales d'une colonne Excel sur une autre colonne.

Usage:
    python copie_fusions.py <fichier> <feuille> <col_exemple> <col_cible> [-o sortie] [--dry-run]

Exemple:
    python copie_fusions.py planning.xlsx "Feuil1" B G
    -> la colonne G reçoit exactement les mêmes fusions verticales que la colonne B

Dépendance : pip install openpyxl
"""

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


def colonne_vers_index(valeur):
    """Accepte une lettre ("B", "ag") ou un numéro ("2")."""
    valeur = str(valeur).strip()
    return int(valeur) if valeur.isdigit() else column_index_from_string(valeur.upper())


def copier_fusions(fichier, feuille, col_exemple, col_cible, sortie=None, dry_run=False):
    fichier = Path(fichier)
    wb = load_workbook(fichier, keep_vba=fichier.suffix.lower() == ".xlsm")

    if feuille not in wb.sheetnames:
        raise SystemExit(f"Onglet introuvable : {feuille!r}\nDisponibles : {wb.sheetnames}")
    ws = wb[feuille]

    src = colonne_vers_index(col_exemple)
    dst = colonne_vers_index(col_cible)
    if src == dst:
        raise SystemExit("La colonne exemple et la colonne cible sont identiques.")

    # 1. Fusions VERTICALES de la colonne modèle : 1 seule colonne de large, > 1 ligne
    modele = sorted(
        (r.min_row, r.max_row)
        for r in ws.merged_cells.ranges
        if r.min_col == r.max_col == src and r.max_row > r.min_row
    )

    # 2. Toute fusion existante touchant la colonne cible doit sauter, sinon conflit
    a_defusionner = [str(r) for r in ws.merged_cells.ranges if r.min_col <= dst <= r.max_col]

    lettre_dst = get_column_letter(dst)
    print(f"Feuille  : {ws.title}")
    print(f"Modèle   : colonne {get_column_letter(src)} -> {len(modele)} fusion(s) verticale(s)")
    for debut, fin in modele:
        print(f"           {lettre_dst}{debut}:{lettre_dst}{fin}")
    if a_defusionner:
        print(f"Défusion : {', '.join(a_defusionner)}")

    if dry_run:
        print("\n--dry-run : rien n'a été enregistré.")
        return

    for plage in a_defusionner:
        ws.unmerge_cells(plage)

    for debut, fin in modele:
        ws.merge_cells(start_row=debut, start_column=dst, end_row=fin, end_column=dst)

    sortie = Path(sortie) if sortie else fichier
    wb.save(sortie)
    print(f"\nEnregistré : {sortie}")


def main():
    p = argparse.ArgumentParser(description="Copie les fusions verticales d'une colonne Excel vers une autre.")
    p.add_argument("fichier", help="chemin du .xlsx / .xlsm")
    p.add_argument("feuille", help="nom de l'onglet")
    p.add_argument("col_exemple", help="colonne modèle, ex. B")
    p.add_argument("col_cible", help="colonne à fusionner, ex. G")
    p.add_argument("-o", "--sortie", help="fichier de sortie (défaut : écrase l'original)")
    p.add_argument("--dry-run", action="store_true", help="affiche sans enregistrer")
    a = p.parse_args()
    copier_fusions(a.fichier, a.feuille, a.col_exemple, a.col_cible, a.sortie, a.dry_run)


if __name__ == "__main__":
    main()
  
