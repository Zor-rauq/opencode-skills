#!/usr/bin/env python3
# /// script
# requires-python = ">=3.9"
# dependencies = ["openpyxl>=3.1"]
# ///
"""
Met a jour un fichier Excel de suivi d'avancement SSS depuis une instance GitLab.

Usage minimal :
    export GITLAB_URL=https://gitlab.interne.local
    export GITLAB_TOKEN=glpat-xxxxxxxxxxxx
    python maj_sss.py --target 1234        --file suivi.xlsx
    python maj_sss.py --target grp/mgmt    --file suivi.xlsx
    python maj_sss.py --target 1234 --file suivi.xlsx --dry-run

Le script est proprietaire de DEUX onglets, qu'il efface et reecrit a chaque
passage :
    - US_GitLab : une ligne par couple (SSS, US). Donnees brutes.
    - Matrice   : une ligne par SSS. Formules de calcul.
Il ne fait qu'AJOUTER des lignes a l'onglet Config (nouveaux statuts decouverts),
et ne touche a AUCUN autre onglet du classeur.
"""

import argparse
import json
import os
import re
import shutil
import ssl
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# Parametres d'extraction -- c'est ici qu'on adapte au format des descriptions
# --------------------------------------------------------------------------

# Identifiant SSS reconnu dans la description. Surchargeable par --sss-regex.
SSS_REGEX = r"\bSSS[-_ ]?[A-Za-z0-9]+(?:[-_.][A-Za-z0-9]+)*"

# Une SSS n'est retenue que sur une ligne de liste a puce markdown ("- ", "* ").
# --sss-anywhere desactive cette restriction et scanne toute la description.
BULLET_RE = re.compile(r"^\s*(?:[-*+]|\d+[.)])\s+")

# Mapping par defaut ecrit dans l'onglet Config a la creation.
# Un poids vide = ligne ignoree dans la moyenne (et non comptee a zero).
CONFIG_DEFAUT = [
    ("opened", 0.30, "Etat GitLab (avant passage aux statuts)"),
    ("in progress", 0.60, "Etat arbitraire hors GitLab / statut 'In progress'"),
    ("closed", 1.00, "Etat GitLab"),
    ("To do", 0.30, "Statut systeme GitLab"),
    ("Done", 1.00, "Statut systeme GitLab"),
    ("Won't do", None, "Poids vide = exclu de la moyenne. A decider."),
    ("Duplicate", None, "Poids vide = exclu de la moyenne. A decider."),
]

ONGLET_CONFIG = "Config"
ONGLET_US = "US_GitLab"
ONGLET_MATRICE = "Matrice"

POLICE = "Arial"


# --------------------------------------------------------------------------
# Client GitLab (stdlib uniquement)
# --------------------------------------------------------------------------


class GitLabError(RuntimeError):
    pass


class Client:
    def __init__(self, base_url, token, insecure=False, timeout=60):
        self.base = base_url.rstrip("/")
        self.token = token
        self.timeout = timeout
        self.ctx = ssl._create_unverified_context() if insecure else None

    def _open(self, req):
        try:
            return urllib.request.urlopen(req, timeout=self.timeout, context=self.ctx)
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", "replace")[:500]
            raise GitLabError(f"HTTP {exc.code} sur {req.full_url}\n{detail}") from None
        except urllib.error.URLError as exc:
            raise GitLabError(f"Connexion impossible a {self.base} : {exc.reason}") from None

    def rest(self, path):
        req = urllib.request.Request(
            f"{self.base}/api/v4/{path.lstrip('/')}",
            headers={"PRIVATE-TOKEN": self.token},
        )
        with self._open(req) as resp:
            return json.load(resp)

    def graphql(self, query, variables):
        req = urllib.request.Request(
            f"{self.base}/api/graphql",
            data=json.dumps({"query": query, "variables": variables}).encode(),
            headers={
                "PRIVATE-TOKEN": self.token,
                "Content-Type": "application/json",
            },
            method="POST",
        )
        with self._open(req) as resp:
            payload = json.load(resp)
        if payload.get("errors"):
            raise GitLabError("; ".join(e.get("message", "") for e in payload["errors"]))
        return payload["data"]


def resoudre_cible(client, ident):
    """Accepte un ID numerique ou un chemin, projet ou groupe, sans distinction."""
    encode = urllib.parse.quote(str(ident), safe="")
    try:
        p = client.rest(f"projects/{encode}")
        return "project", p["path_with_namespace"], p["name"]
    except GitLabError:
        pass
    try:
        g = client.rest(f"groups/{encode}")
        return "group", g["full_path"], g["name"]
    except GitLabError:
        pass
    raise GitLabError(
        f"'{ident}' ne correspond ni a un projet ni a un groupe accessible avec ce token."
    )


REQUETE = """
query($path: ID!, $after: String%(varLabels)s) {
  %(racine)s(fullPath: $path) {
    workItems(first: 100, after: $after, types: [ISSUE]%(argLabels)s) {
      pageInfo { hasNextPage endCursor }
      nodes {
        iid
        title
        state
        webUrl
        widgets {
          ... on WorkItemWidgetDescription { description }
          ... on WorkItemWidgetLabels { labels { nodes { title } } }
          %(fragStatus)s
        }
      }
    }
  }
}
"""

FRAG_STATUS = "... on WorkItemWidgetStatus { status { name } }"


def _construire_requete(racine, avec_status, avec_filtre_label):
    return REQUETE % {
        "racine": racine,
        "varLabels": ", $labels: [String!]" if avec_filtre_label else "",
        "argLabels": ", labelName: $labels" if avec_filtre_label else "",
        "fragStatus": FRAG_STATUS if avec_status else "",
    }


def recuperer_us(client, racine, path, label):
    """Pagine les work items. Degrade automatiquement si l'instance ne supporte
    pas le widget Status ou le filtre labelName."""
    avec_status, avec_filtre = True, True
    for _ in range(3):
        try:
            return _paginer(client, racine, path, label, avec_status, avec_filtre), avec_status
        except GitLabError as exc:
            msg = str(exc).lower()
            if avec_status and "status" in msg:
                avertir("widget Status indisponible -> repli sur l'etat opened/closed.")
                avec_status = False
                continue
            if avec_filtre and "labelname" in msg:
                avertir("filtre labelName refuse par l'API -> filtrage cote script.")
                avec_filtre = False
                continue
            raise
    raise GitLabError("Impossible de construire une requete acceptee par l'instance.")


def _paginer(client, racine, path, label, avec_status, avec_filtre):
    query = _construire_requete(racine, avec_status, avec_filtre)
    variables = {"path": path, "after": None}
    if avec_filtre:
        variables["labels"] = [label]

    items, curseur = [], None
    while True:
        variables["after"] = curseur
        data = client.graphql(query, variables)
        racine_data = data.get(racine)
        if racine_data is None:
            raise GitLabError(f"'{path}' inaccessible (droits insuffisants ?).")
        bloc = racine_data["workItems"]
        items.extend(bloc["nodes"])
        if not bloc["pageInfo"]["hasNextPage"]:
            break
        curseur = bloc["pageInfo"]["endCursor"]

    # Filtrage cote script : garantit le label meme si le serveur ne filtre pas.
    resultat = []
    for n in items:
        labels = _widget(n, "labels")
        noms = [x["title"] for x in labels["labels"]["nodes"]] if labels else []
        if noms and label.lower() not in [x.lower() for x in noms]:
            continue
        resultat.append(n)
    return resultat


def _widget(node, cle):
    """Retourne le premier widget contenant la cle demandee."""
    for w in node.get("widgets") or []:
        if cle in w:
            return w
    return None


# --------------------------------------------------------------------------
# Extraction des SSS
# --------------------------------------------------------------------------


def extraire_sss(description, motif, partout=False):
    if not description:
        return []
    trouves, vus = [], set()
    for ligne in description.splitlines():
        if not partout and not BULLET_RE.match(ligne):
            continue
        for m in motif.finditer(ligne):
            sss = m.group(0).strip().rstrip(".,;:")
            if sss.upper() not in vus:
                vus.add(sss.upper())
                trouves.append(sss)
    return trouves


def construire_lignes(items, motif, partout, base_url, path):
    """Une ligne par couple (SSS, US). Une US citant 3 SSS produit 3 lignes."""
    lignes, orphelines = [], []
    for n in items:
        desc = _widget(n, "description")
        st = _widget(n, "status")
        statut = (st or {}).get("status") or {}
        sss_list = extraire_sss(desc.get("description") if desc else "", motif, partout)
        commun = {
            "us": f"#{n['iid']}",
            "titre": n["title"],
            "statut": statut.get("name") or "",
            "etat": "opened" if n["state"].upper() == "OPEN" else "closed",
            "url": n.get("webUrl") or f"{base_url}/{path}/-/issues/{n['iid']}",
        }
        if not sss_list:
            orphelines.append(commun)
            continue
        for sss in sss_list:
            lignes.append(dict(commun, sss=sss))
    lignes.sort(key=lambda r: (_cle_naturelle(r["sss"]), _cle_naturelle(r["us"])))
    return lignes, orphelines


def _cle_naturelle(s):
    return [int(x) if x.isdigit() else x.lower() for x in re.split(r"(\d+)", s)]


# --------------------------------------------------------------------------
# Ecriture du classeur
# --------------------------------------------------------------------------

GRAS = Font(name=POLICE, bold=True, color="FFFFFF")
FOND_TITRE = PatternFill("solid", fgColor="44546A")
FOND_ALERTE = PatternFill("solid", fgColor="FFFF00")
NORMAL = Font(name=POLICE)
BLEU = Font(name=POLICE, color="0000FF")


def _entetes(ws, titres, largeurs):
    for i, t in enumerate(titres, start=1):
        c = ws.cell(row=1, column=i, value=t)
        c.font = GRAS
        c.fill = FOND_TITRE
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = largeurs[i - 1]
    ws.freeze_panes = "A2"


def _remplacer_onglet(wb, nom):
    """Efface l'onglet s'il existe et le recree a la meme position."""
    index = wb.sheetnames.index(nom) if nom in wb.sheetnames else len(wb.sheetnames)
    if nom in wb.sheetnames:
        del wb[nom]
    return wb.create_sheet(nom, index)


def assurer_config(wb, statuts_vus):
    """Cree l'onglet Config s'il manque, puis AJOUTE les valeurs inconnues.
    Ne modifie jamais une ligne existante : les poids saisis sont preserves."""
    nouveau = ONGLET_CONFIG not in wb.sheetnames
    if nouveau:
        ws = wb.create_sheet(ONGLET_CONFIG, 0)
        _entetes(ws, ["Valeur (statut ou etat)", "Poids", "Commentaire"], [30, 10, 55])
        for valeur, poids, note in CONFIG_DEFAUT:
            _ajouter_config(ws, valeur, poids, note)
        ws["E2"] = "A MODIFIER : colonne B uniquement. Un poids vide exclut la ligne du calcul."
        ws["E3"] = "La correspondance est insensible a la casse ('closed' = 'Closed')."
        ws["E4"] = "Les nouveaux statuts detectes sont ajoutes automatiquement, poids vide."
        for ref in ("E2", "E3", "E4"):
            ws[ref].font = Font(name=POLICE, size=9, italic=True, color="808080")
        ws.column_dimensions["E"].width = 70
    else:
        ws = wb[ONGLET_CONFIG]

    connues = {
        str(ws.cell(row=r, column=1).value).strip().lower()
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=1).value
    }
    ajoutes = []
    for valeur in sorted(statuts_vus):
        if valeur and valeur.strip().lower() not in connues:
            _ajouter_config(ws, valeur, None, "Detecte automatiquement -- poids a saisir")
            connues.add(valeur.strip().lower())
            ajoutes.append(valeur)
    return ws.max_row, ajoutes


def _ajouter_config(ws, valeur, poids, note):
    r = max(2, ws.max_row + 1)
    ws.cell(row=r, column=1, value=valeur).font = NORMAL
    c = ws.cell(row=r, column=2, value=poids)
    c.font = BLEU
    c.number_format = "0%"
    if poids is None:
        c.fill = FOND_ALERTE
    ws.cell(row=r, column=3, value=note).font = Font(
        name=POLICE, size=9, italic=True, color="808080"
    )


def ecrire_us(wb, lignes, n_config, horodatage):
    ws = _remplacer_onglet(wb, ONGLET_US)
    _entetes(
        ws,
        ["SSS", "US", "Titre", "Statut", "Etat", "Avancement retenu", "Poids", "URL", "MAJ"],
        [16, 9, 55, 18, 11, 20, 10, 46, 18],
    )
    fin_cfg = max(2, n_config)
    for i, l in enumerate(lignes, start=2):
        ws.cell(row=i, column=1, value=l["sss"])
        ws.cell(row=i, column=2, value=l["us"])
        ws.cell(row=i, column=3, value=l["titre"])
        ws.cell(row=i, column=4, value=l["statut"])
        ws.cell(row=i, column=5, value=l["etat"])
        # Le statut prime sur l'etat ; la logique reste lisible dans la cellule.
        ws.cell(row=i, column=6, value=f'=IF($D{i}<>"",$D{i},$E{i})')
        # Double INDEX volontaire : MATCH peut reussir sur une valeur dont le
        # poids n'est pas encore saisi. INDEX renvoie alors 0, qui serait moyenne
        # comme un 0 % legitime. On force le vide pour exclure la ligne.
        lookup = (
            f"INDEX({ONGLET_CONFIG}!$B$2:$B${fin_cfg},"
            f"MATCH($F{i},{ONGLET_CONFIG}!$A$2:$A${fin_cfg},0))"
        )
        ws.cell(
            row=i,
            column=7,
            value=f'=IFERROR(IF({lookup}="","",{lookup}),"")',
        ).number_format = "0%"
        ws.cell(row=i, column=8, value=l["url"])
        ws.cell(row=i, column=9, value=horodatage)
        for col in range(1, 10):
            ws.cell(row=i, column=col).font = NORMAL
    ws.auto_filter.ref = f"A1:I{max(2, len(lignes) + 1)}"
    return max(2, len(lignes) + 1)


def ecrire_matrice(wb, lignes, fin_us):
    ws = _remplacer_onglet(wb, ONGLET_MATRICE)
    _entetes(ws, ["SSS", "Avancement", "Nb US", "Dont non mappees"], [18, 14, 10, 18])
    sss_uniques = sorted({l["sss"] for l in lignes}, key=_cle_naturelle)
    for i, sss in enumerate(sss_uniques, start=2):
        ws.cell(row=i, column=1, value=sss).font = NORMAL
        ws.cell(
            row=i,
            column=2,
            value=f'=IFERROR(AVERAGEIF({ONGLET_US}!$A$2:$A${fin_us},$A{i},'
            f'{ONGLET_US}!$G$2:$G${fin_us}),"")',
        ).number_format = "0.0%"
        ws.cell(row=i, column=3, value=f"=COUNTIF({ONGLET_US}!$A$2:$A${fin_us},$A{i})")
        ws.cell(
            row=i,
            column=4,
            value=f"=SUMPRODUCT(({ONGLET_US}!$A$2:$A${fin_us}=$A{i})*"
            f'({ONGLET_US}!$G$2:$G${fin_us}=""))',
        )
        for col in (2, 3, 4):
            ws.cell(row=i, column=col).font = NORMAL
            ws.cell(row=i, column=col).alignment = Alignment(horizontal="center")
    return len(sss_uniques)


# --------------------------------------------------------------------------


def avertir(msg):
    print(f"  [!] {msg}", file=sys.stderr)


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--target", required=True, help="ID ou chemin du projet ou du groupe")
    p.add_argument("--file", required=True, help="Fichier .xlsx a mettre a jour (cree si absent)")
    p.add_argument("--url", default=os.environ.get("GITLAB_URL"), help="URL de l'instance (ou $GITLAB_URL)")
    p.add_argument("--token", default=os.environ.get("GITLAB_TOKEN"), help="Token API (ou $GITLAB_TOKEN)")
    p.add_argument("--label", default="US", help="Label identifiant les US (defaut : US)")
    p.add_argument("--sss-regex", default=SSS_REGEX, help="Motif d'identifiant SSS")
    p.add_argument("--sss-anywhere", action="store_true", help="Chercher hors des listes a puces")
    p.add_argument("--insecure", action="store_true", help="Ignorer la validation du certificat TLS")
    p.add_argument("--no-backup", action="store_true", help="Ne pas creer de .bak")
    p.add_argument("--dry-run", action="store_true", help="Afficher sans ecrire")
    args = p.parse_args()

    if not args.url or not args.token:
        p.error("--url et --token sont requis (ou $GITLAB_URL / $GITLAB_TOKEN)")

    client = Client(args.url, args.token, insecure=args.insecure)
    racine, path, nom = resoudre_cible(client, args.target)
    print(f"Cible   : {racine} '{nom}' ({path})")

    items, avec_status = recuperer_us(client, racine, path, args.label)
    print(f"Issues  : {len(items)} portant le label '{args.label}'"
          f"{'' if avec_status else ' (statuts indisponibles)'}")

    motif = re.compile(args.sss_regex, re.IGNORECASE)
    lignes, orphelines = construire_lignes(items, motif, args.sss_anywhere, args.url, path)
    sss_uniques = {l["sss"] for l in lignes}
    print(f"Extrait : {len(lignes)} couples (SSS, US) sur {len(sss_uniques)} SSS")

    if orphelines:
        avertir(f"{len(orphelines)} US sans SSS dans la description : "
                + ", ".join(o["us"] for o in orphelines[:15])
                + (" ..." if len(orphelines) > 15 else ""))

    if args.dry_run:
        for l in lignes[:40]:
            print(f"    {l['sss']:<14} {l['us']:<7} {l['statut'] or l['etat']:<14} {l['titre'][:50]}")
        if len(lignes) > 40:
            print(f"    ... et {len(lignes) - 40} autres lignes")
        return 0

    existe = os.path.exists(args.file)
    if existe and not args.no_backup:
        shutil.copy2(args.file, args.file + ".bak")
    wb = load_workbook(args.file) if existe else Workbook()
    if not existe:
        wb.remove(wb.active)

    horodatage = datetime.now().strftime("%Y-%m-%d %H:%M")
    statuts_vus = {l["statut"] or l["etat"] for l in lignes}
    n_config, ajoutes = assurer_config(wb, statuts_vus)
    fin_us = ecrire_us(wb, lignes, n_config, horodatage)
    n_sss = ecrire_matrice(wb, lignes, fin_us)

    wb.calculation.fullCalcOnLoad = True
    tmp = args.file + ".tmp"
    wb.save(tmp)
    os.replace(tmp, args.file)

    print(f"Ecrit   : {ONGLET_US} ({len(lignes)} lignes), {ONGLET_MATRICE} ({n_sss} SSS)")
    if ajoutes:
        avertir("Nouvelles valeurs ajoutees dans Config, poids a saisir : " + ", ".join(ajoutes))
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except GitLabError as exc:
        print(f"Erreur GitLab : {exc}", file=sys.stderr)
        sys.exit(1)
    except KeyboardInterrupt:
        sys.exit(130)
      
